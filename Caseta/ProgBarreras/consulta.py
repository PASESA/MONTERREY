import openpyxl
import pymysql
 
 
class ExcelUtils4Others(object):
 
    def __init__(self, filename):
        self.wb = openpyxl.load_workbook(filename)
        self.wses = self.wb.sheetnames
        self.ws = self.wb.active
 
    def read_sheet(self, sheet_name=""):
        if sheet_name:
            ws = self.wb[sheet_name]
        else:
            ws = self.ws
        for row in ws.rows:
            for cell in row:
                print(cell.value)
 
    @classmethod
            # ~ conn=pymysql.connect(host="localhost",
                                 # ~ user="Aurelio",
                                 # ~ passwd="RG980320",
                                 # ~ database="Parqueadero1")

        # ~ #conexion = pymysql.connect(host="192.168.1.91",
        # ~ #                   user="Aurelio",
        # ~ #                   passwd="RG980320",
        # ~ #                   database="Parqueadero1")
        # ~ return conexion
    def get_conn(self, db, user='Aurelio', pwd='RG980320', host='localhost', charset='utf8'):
        """
                 : param db: la base de datos a conectar
                 : param user: nombre de usuario de la base de datos
                 : param host: dirección de host de la base de datos (ip)
                 : param pwd: contraseña de la base de datos
        :param charset: utf8
        :return:
        """
        try:
            conn = pymysql.connect(
                db=Parqueadero1,
                user=user,
                host=host,
                password=pwd,
                charset=charset
            )
        except:
            pass
        return conn
 
    def export_xls(self, db, sheet ='', paras=None):
                 "" "Exportar desde la base de datos mysql a Excel" ""
       # global sql
        if paras is None:
            paras = []
        conn = ExcelUtils4Others.get_conn()
        cursor = conn.cursor()
        for para in paras:
            sql = "SELECT `" + para + "`, "
        sql += "FROM " + db
        index = sql.rfind(',')
        sql = sql[:index] + sql[index+1:]
        print(sql)
 
        cursor.execute(sql)
        rows = cursor.fetchall()
        for(i, row) in enumerate(rows):
            self.ws['A{0}'.format(i+1)] = row[0]
            self.ws['B{0}'.format(i+1)] = row[0]
            self.ws['C{0}'.format(i+1)] = row[0]
